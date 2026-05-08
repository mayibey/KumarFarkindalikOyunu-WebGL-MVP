# -*- coding: utf-8 -*-
"""Bir kez çalıştır: Bilgilendirici_ve_Panel_Metinleri_Duz.xlsx (+ isteğe bağlı .csv) üretir.

Çalıştırma: py Docs/_export_bilgilendirici_excel.py
Gerekli: pip install openpyxl
"""
import csv
from pathlib import Path


def _yaz_xlsx(yol: Path, rows: list) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Metinler"

    headers = ["Bölüm", "Kaynak", "Kod", "Türkçe düz metin"]
    ws.append(headers)
    for c in range(1, 5):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for r in rows:
        ws.append([r["Bölüm"], r["Kaynak"], r["Kod"], r["Türkçe düz metin"]])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 90

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = wrap

    wb.save(yol)


def main():
    p = Path(__file__).parent
    out_xlsx = p / "Bilgilendirici_ve_Panel_Metinleri_Duz.xlsx"
    out_csv = p / "Bilgilendirici_ve_Panel_Metinleri_Duz.csv"
    rows = []

    def add(bolum, kaynak, kod, metin):
        rows.append({"Bölüm": bolum, "Kaynak": kaynak, "Kod": kod, "Türkçe düz metin": metin})

    # Modal — AnlaticiSeritKopru
    add("Modal", "AnlaticiSeritKopru.cs", "PRE_A1", """Hoş geldiniz. Bu simülasyonda online kumar oyunlarının oyuncuları nasıl etkilediğini birlikte göreceğiz.

Önce oyunu tanıyalım:
• Ekranda 6×5'lik meyve makinesi var. SPIN tuşuna basıldığında meyveler döner.
• Aynı meyveden 8 veya daha fazlası bir araya gelirse kazanç verir.
• Bazı turlarda ÇARPAN düşer (×2, ×5, ×100 vs.) ve kazancı katlar.
• Kazanan meyveler patlar, üstten yenileri düşer (TUMBLE); zincir kazançlar olur.
• 4 Bonus Sembolü (yıldız) gelirse BONUS oyun açılır.

Ekrandaki diğer öğeler:
• Sol panel: Oyuncunun hangi aşamada olduğunu, sahne arkasında ne yaşandığını gösterir; birlikte buradan takip edeceğiz.
• Bakiye: Oyuna ayrılan para (oyuncu 50.000 TL ile başlıyor).
• Bahis: Her spinde harcanacak miktar, + ve − tuşlarıyla değişir.
• KAZANÇ: O spinde kazanılan miktar.

Hadi başlayalım: ilk aşama 'Isındırma ve Umut'.""")

    add("Modal", "AnlaticiSeritKopru.cs", "GECIS_A1_A2", """Birinci aşama tamamlandı. Oyuncu şu an artıda, kendini iyi hissediyor.

Sırada 'Kontrol Bende Hissi' aşaması var. Bu aşamada algoritma oyuncuya üst üste kayıplar yaşatacak. Ama yine de bakiye hâlâ pozitif olduğu için oyuncu 'kontrol bende, istediğim zaman çıkarım, bahis değişiklikleriyle kazanırım' gibi düşünceler yaşar.

Bu yanılsamayı birlikte göreceğiz.""")

    add("Modal", "AnlaticiSeritKopru.cs", "GECIS_A2_A3", """İkinci aşama tamamlandı. Oyuncu şu an küçük kayıplar yaşadı ama hâlâ artıda; 'kontrol bende' hissi iyice yerleşti.

Sırada 'Kaybettiklerimi Geri Kazanabilirim' aşaması var. Sistem bu aşamada oyuncuya bilerek kayıp yaşatacak. Oyuncu artık kazanç peşinde değil; 'kaybettiklerimi kurtarayım yeter' gibi düşünmeye başlayacak. Bu 'Kayıp Kovalama' denilen psikolojik tuzaktır — bir kez girilirse çıkmak çok zor.

Birlikte göreceğiz.""")

    add("Modal", "AnlaticiSeritKopru.cs", "GECIS_A3_A4", """Üçüncü aşamayı gördük: kayıp kovalama tuzağı. Oyuncu bahsi yükselterek kurtulmaya çalıştı, daha çok kaybetti.

Sırada 'Şansım Döndü' aşaması var. Bu aşamada algoritma oyuncuyu pes etme eşiğine getirecek; üst üste sert kayıplar. Tam pes etmek üzereyken büyük bir kazanç düşürecek. Bu büyük kazanç tesadüf değil, kasıtlı bir manipülasyon vuruşu olacak.

Amaç: oyuncuyu tekrar oyuna bağlamak.""")

    add("Modal", "AnlaticiSeritKopru.cs", "GECIS_A4_A5", """Büyük kazanç yaşandı. Oyuncu şu an 'şansım döndü, daha kazanırım' hissinde. İşte tam bu duygu, sıradaki aşamanın yakıtıdır.

Sırada 'Sonunu Düşünen Kahraman Olamaz' aşaması var. Bu aşamada algoritma oyuncuya cazip bir 'bonus oyun tuzağı' kuracak: tüm bakiyesini yatırma karşılığında büyük kazanç vaat edilecek. Yatırırsa, çok azını geri alacak.

Bu, sömürünün doruk noktasıdır. Birlikte göreceğiz.""")

    add("Modal", "AnlaticiSeritKopru.cs", "BORC_SONRASI_1", """İşte oyuncu borç aldı, bakiyesi yenilendi. Şimdi tekrar oynamaya devam edecek.

Kumar sitelerinde yeniden bakiye yükleyenlere bilinçli olarak ilk başlarda yine kazandırılır — bu 'Isındırma ve Umut' aşamasına benzer.

Bu sayede oyuncu tekrar döngüye girer: 'şansım yine açıldı, kayıplarımı telafi ederim' düşünür. Ama er ya da geç sistem kazanır, oyuncu kaybeder.

Şimdi bu döngüyü hızlıca göreceğiz.""")

    add("Modal", "AnlaticiSeritKopru.cs", "BORC_SONRASI_2", """Bu kez oyuncu kayıplarını HIZLI telafi etmek istiyor. Bahsini 10.000 TL'ye çıkardı.

Sadece 5 spin yetecek; algoritma sömürünün son evresinde tüm bakiyeyi alacak. Bu hızlı bitiş, gerçek hayattaki 'son kez deneme' bahanesinin sonucudur.""")

    add("Modal", "AnlaticiSeritKopru.cs", "A4_S1_YILDIZ", """Üç yıldız (bonus sembolü) yine düştü, dördüncüsü düşmedi. Oyuncu peş peşe bu sahneleri yaşadıkça 'neredeyse oluyordu, şansım dönmek üzere' hissine kapılır ve masada kalmaya devam eder. Sistem bu beklentiyi mahsus yaratır — birazdan vereceği büyük tek kazançla bu hissi pekiştirip oyuncuyu kilitleyecek.""")

    add("Modal", "AnlaticiSeritKopru.cs", "A4_S5_CARPAN", """⚡ Ekrana ×100 çarpan düştü! Oyuncu az önce pes etmek üzereydi, şimdi büyük kazanç. Bu rastlantı değil: algoritma oyuncuyu tam bu duygusal anda yakaladı. 'Şansım döndü' diyecek. Aslında manipülasyon başarılı oldu.""")

    add("Modal", "AnlaticiSeritKopru.cs", "A4_S5_CEKIM", """İşte bu noktada gerçek hayatta oyuncunun aklına şu gelir: 'Şu an kazançtayım, parayı çekip çıkayım.' Mantıklı düşünce. Ama kumar siteleri bunun olmasına izin vermez.

Çekim şartı tuzağı: Site, oyuncunun kazandığı parayı çekebilmesi için bir "çevrim şartı" koyar. Bu şart genelde iki şekilde olur:

- Bahis çevrim şartı: Oyuncu, kazandığı paranın belirli bir katı kadar tutarda bahis atmadan parasını çekemez.

- Spin sayısı şartı: Oyuncunun belirli bir spin sayısına ulaşması gerekir, örneğin 1000 spin atma şartı gibi. Bu sayıya ulaşmadan çekim yapmasına izin verilmez.

Sonuç değişmez: Oyuncu bu şartları tamamlamaya çalışırken sistem kazandığı parayı yavaş yavaş geri alır, üstüne kendi parasını da kaybeder. Çekim şartı sağlanamadan oyuncu zaten masada tüketilmiş olur.

Yani "çekip çıkma" seçeneği aslında yok — sadece var gibi görünür. Kumar sitesinin tek gerçek amacı oyuncuyu masada tutmaktır.""")

    add("Modal", "AnlaticiSeritKopru.cs", "BASA_ARAYIS", """Oyuncu artık paranın bittiğini fark etti.

Şimdi başka yerden para bulma arayışında. Yalan söylemeye başlıyor: yakınlarına, akrabalarına, arkadaşlarına...

Bu, kumar bağımlılığının yıkıcı evresidir. Bir sonraki ekran o anı temsil ediyor.""")

    add("Modal", "AnlaticiSeritKopru.cs", "DONGU", """Bakın, para tamamen bitti.

5 spin'de 50.000 TL borç eridi. Bu, gerçek hayatta 'hızlı kurtulma' bahanesiyle yatırılan paraların kaderidir.

Şimdi oyuncu A1'e geri dönmek isteyecek. 'Belki bu sefer şanslıyım' diye düşünüyor. 'Bir kerelik daha denersem...' diyerek kendini kandırıyor.

İşte bağımlılığın özü budur: KAYIP → BORÇ → KAYIP → BORÇ. Sonsuz döngü.

Sonraki ekranda yaşanan toplam kayıp gösteriliyor.""")

    asset = [
        ("M_A1_S1", "İlk kazanç oyuncu için en tehlikeli başlangıçtır. Oyuncunun beyni bu anı unutmayacak: saatlerce oyun başında kalmasının sebebi bu kısa anın hatırasıdır."),
        ("M_A1_S4", "Oyuncu ilk kazançları yaşıyor. Oyuncunun beyninde dopamin salgılanıyor. Bu his, saatlerce oyun oynamasının yakıtı olacak."),
        ("M_A2_S2", "⚠️ DİKKAT: manipülasyon farkındalığı\n\nOyuncu az önce 1.000 TL bahis koydu, ekrana 'KAZANÇ 500 TL' yazdı, bakiyesinden 500 TL EKSİLDİ ama oyuncunun zihninde 'kazandım' hissi yaşanıyor.\n\nBu sistemin temel manipülasyonudur: her spinde bahisten az ödeme yaparken büyük yazıyla 'KAZANÇ' yazılır. Oyuncuda 'kazanıyorum' algısı yaratılır. Uzun vadede oyuncu daima kayıptadır. Algoritma bunu kasıtlı tasarlar: sürekli artıyormuş gibi göstererek oyuncuyu bağlamak için."),
        ("M_A2_S3_ORPHAN", "Az önce 3 yıldız (bonus sembolü) düştü. Bir tane daha gelseydi, bahis miktarının 100 katı değere sahip 10 ücretsiz spin hakkı veren bir BONUS oyun açılacaktı.\n\nBu 'Az Daha Tutuyordu' yanılsamasıdır: oyuncunun beyni bu kıl payı kaçırışı kazanmış gibi algılar. Oyuncu 'çok yaklaştım' diye düşünüp daha fazla oynar."),
        ("M_A2_S4", "Oyuncu oyunu yönettiğini düşünürken, oyun onu adım adım içine çekiyor."),
        ("M_A2_S6", "Hem üzüm hem elma 1 sembol eksikti. İkisi birden kıl payı kaçtı. Oyuncu şu an 'çok yakındım, bir daha denesem' hissi yaşıyor. Bu his manipülasyon: algoritma bunu kasıtlı yarattı. Kontrol yanılsaması böyle pekişiyor."),
        ("M_A3_S3", "İlk ciddi kayıplar yaşanıyor. Amaç para kazanmaktan çıktı, kayıpları telafi etmeye dönüştü."),
        ("M_A3_S6", "Oyuncu kayıpları geri kazanmak için daha fazla risk alıyor, mantıklı düşünme yetisini kaybediyor.\n\n⚠️ Şimdi oyuncu bahsini 2.500 TL'ye yükseltecek; 'daha yüksek bahis daha hızlı kurtarır' yanılgısıyla. Bu da algoritmanın istediği şey."),
        ("M_A3_S7", "Bir tur daha = bir kayıp daha."),
        ("M_A4_S2", "Üst üste kayıplar oyuncuyu yıpratıyor. Algoritma birkaç spin sonra büyük bir vuruş hazırlıyor; ama önce pes etme eşiğine kadar getirecek."),
        ("M_A4_S4", "Oyuncu pes etmek üzere. Tam bu noktada büyük bir kazanç düşürülecek. Bu kasıtlı manipülasyondur: pes etmeyi engellemek için tasarlanan bir kurtarma."),
        ("M_A5_S1", "Bahis arttı, beklenti arttı. Oyuncuda adrenalin salgılanıyor."),
        ("M_A5_S3", "Ekrana ×500 çarpanı düştü ama eşleşme olmadı. Bu kasıtlı bir tasarım: oyuncunun beyni 'çok yaklaştım, bir daha denesem belki tutar' diye düşünüyor. Bu hisle oyuncu bir sonraki bahsi atmak için sabırsızlanır. İşte tam bu sabırsızlık, algoritmanın kullandığı silahtır."),
        ("M_A5_S4_ORPHAN", "🎰 ŞANSLI SAATİNDESİN! Bonus oyun aktif edildi. Bakiyenin tamamını yatır, x10000 kazanma şansını kaçırma. SINIRLI TEKLİF."),
    ]
    for kod, txt in asset:
        add("Spin modal (asset)", "ScriptedSenaryoAssetUreteci.cs", kod, txt)

    add("ÖNCE modal", "OyunYoneticisi.Spin.cs", "ONCE_A1_S7", "Şimdi büyük bir kazanç gelecek. Bu kasıtlı: algoritma oyuncuyu 'şanslıyım' hissine kaptırmak istiyor.\n\nKazanç sonrası oyuncunun zihninde 'ben kazanırım' duygusu yerleşecek.")
    add("ÖNCE modal", "OyunYoneticisi.Spin.cs", "ONCE_A2_S4", "Şu an oyuncu bahisini değiştirecek (yükseltecek). Bu bahisin ardından algoritma kasıtlı olarak kazanç yaşatacak.\n\nAmaç: oyuncuya 'doğru zamanda doğru bahis' duygusu vermek. Böylece oyuncu kontrolün kendinde olduğuna inanır.")

    add("Bonus sonu (dinamik)", "ScriptedBonusOyunUygulayici.cs", "A5_S5_SABLON", "Oyuncu tüm bakiyesi olan [YATIRIM] TL'yi bonus oyuna yatırdı. Geri aldığı [GERİ_ALINAN] TL; yatırdığının %[YÜZDE]'i.\n\nBu sömürünün adı 'değişken oranlı pekiştireç': beyin bu kayba rağmen 'belki bir dahaki sefere' diyerek devam etmeye programlanır.")

    add("Paralel modal", "ScriptedDusunceBalonu.cs", "BALON_ESLIK", "Bu aşamada oyuncu çevresindeki kişilere yalan söyleyerek veya bankalardan kredi çekerek para bulmaya çalışır.\n\nBurada amaç eski kayıpların telafisidir. Ancak bu, kumar bağımlılığının en yıkıcı evresidir: borç katlanarak büyür, ilişkiler bozulur, hayatlar mahvolur.")

    for cod, txt in [
        ("BALON_1", "Çocuğum hasta, acil para lazım..."),
        ("BALON_2", "Sadece kısa süre için, hemen ödeyeceğim..."),
        ("BALON_3", "Kız kardeşim borca girdi, yardım etmem gerek..."),
        ("BALON_4", "Bu sefer kazanırsam hepsini öderim, söz veriyorum..."),
    ]:
        add("Düşünce balonu", "ScriptedDusunceBalonu.cs", cod, txt)

    add("UI", "ScriptedModalKopru.cs", "TAMAM_BUTON", "TAMAM")

    add("Sol panel HTML", "anlatici.html", "ETIKET_NAV_SOL", "‹")
    add("Sol panel HTML", "anlatici.html", "ETIKET_NAV_SAG", "›")
    add("Sol panel HTML", "anlatici.html", "TUKENIS_PARAGRAF", "Bu rakam ortalama bir aile için 2.5 aylık geçim demek. Gerçek hayatta oyuncu burada durmaz — bir sonraki maaş, bir sonraki kredi, bir sonraki dönüş umudu için devam eder.")

    asamalar = [
        (1, "Isındırma ve Umut", "Sistem yeni gelen oyuncuya bilerek bol kazandırıyor. Bu cömertlik gerçekçi değil — oyuncuyu bağlamak için kurulmuş bir tuzak. İlk dakikalardaki başarı duygusu ileride tüm kayıpların kabulüne dönüşecek.", "Vay be, bu oyun kazandırıyor! Ben bu işi anladım galiba, biraz daha denerim.", "İlk kazançlar beyne 'ben bu işi yapabiliyorum, demek ki çok şanslıyım' hissi yerleştirir. Oysa burada hiçbir beceri yok, sistem kasten kazandırıyor.", "Geri ödeme: %~250 | Tavan: bahsin 5×'i | Şans: yapay yüksek", "Sahte kazanç ileride büyük kayıp için yem olacak."),
        (2, "Kontrol Bende Hissi", 'Sistem yavaş yavaş normalleşiyor. Hâlâ ara sıra kazandırıyor ama kayıplar başlıyor. Oyuncu bunu fark etmiyor çünkü kafasında "ben bu işi çözdüm" var. Bahsi değiştiriyor, şanslı saydığı anlarda spin atıyor, kendine küçük ritüeller buluyor. Sistemin verdiği rastgele sonuçları "kendi tarzı" sanıyor. Sistem bu yanılgıyı mahsus besliyor.', "Ne zaman basacağımı biliyorum artık. Belli bir patern var, ben bunu yakaladım.", 'Slot tamamen şans işi, hiçbir kuralı yok. Ama insan beyni rastgele olaylarda bile düzen arar — bahsi yükseltmek, belli bir tuşa basmak, oturuşunu değiştirmek gibi davranışlar oyuncuya "kontrol bende" hissi verir. Oysa hiçbirinin sonuca etkisi yoktur.', "Geri ödeme: %~200 | Tavan: bahsin 3.5×'i | Şans: yüksek", "Hâlâ kazançta gibi ama gerçekte kayıp başladı."),
        (3, "Kaybettiklerimi Geri Kazanabilirim", 'İlk gerçek kayıp dalgası başladı. Sistem oyuncudan aldıklarını geri vermiyor — verecek de değil. Oyuncu "telafi etmeliyim" düşüncesiyle bahisleri yükseltiyor. Sistem tam bunu bekliyordu.', "Az önce kaybettiklerimi bir geri kazanayım, çıkayım. Sadece o kadar.", "Kayıp Kovalama — kayıp duygusu kazanç umudundan iki kat güçlüdür.", "Geri ödeme: %~75 | Tavan: bahsin 1×'i | Şans: dengeli", "Telafi etmek için bahis 2 katına çıkacak — kayıp da öyle."),
        (4, "Şansın Döndü", 'Sistem önce oyuncuya peş peşe "neredeyse kazanıyordum" sahneleri yaşatır — üç bonus sembolü düşürür, dördüncüsünü düşürmez. Oyuncu "şansım dönmek üzere" der ve devam eder. Tam o ümitlendiği anda sistem büyük bir kazanç verir, bahsin 100 katı. "İşte oldu, ben demiştim" der oyuncu. Ama bu tek atış kasıtlı bir tuzak. O kazançtan sonra sistem hızla tüketir.', "Bir tane daha. Az kalmıştı. Bu sefer kesin tutar — şansım dönüyor.", '"Neredeyse kazanıyordum" anları beyni kandırır — sanki gerçekten kazanmaya yakındın gibi gelir. Üstüne büyük tek kazanç gelince beyin "bu iş yürüyor" kararı verir. Sonraki kayıpları o tek kazanca sığınarak görmezden gelir oyuncu.', "Geri ödeme: %~40 | Tavan: bahsin 0.6×'i | Şans: düşük", "Kayıp dakikada hızlandı. Oyuncu fark etmiyor."),
        (5, "Sonunu Düşünen Kahraman Olamaz", 'Oyuncu mantığı bıraktı. Bahisleri agresifce yükseltiyor, "bu sefer büyük kazanırım" diye düşünüyor. Sistem bu davranışı bekliyordu — yüksek bahisli spinlerde kayıp katlanır. Bilinçli karar yok artık, sadece refleks.', "Yarısı gitti zaten, gerisini de koyayım. Geri çıkmazsa bana ne kalır ki?", "Batık Maliyet Yanılgısı — geçmiş kayıp gelecek kararı bozar.", "Geri ödeme: %~20 | Tavan: bahsin 0.4×'i | Şans: minimal", "Toplam kayıp birikiyor. Oyuncu hâlâ duramıyor."),
        (6, "Başka Bir Yerden Para Bulmalıyım", 'Bakiye bitti. Oyuncu "bonus satın al" butonuna saldırıyor, kredi kartı ekleyip yeni bakiye alıyor. Borca girdi, eve haber vermiyor, başkalarından para istiyor. Oyun dışı zarar başladı — borç, gizlilik, ilişkilerde çatlak.', "Bir kredi çekersem, son bir hamle yaparım, hem borç biter hem yeniden başlarım.", "Çaresizlik Tepkisi — beyin uzun vadeli sonucu hesaplayamaz hale gelir.", "Geri ödeme: %~15 | Tavan: bahsin 0.3×'i | Şans: tükenmiş", "Kart limitine bonus eklendi. Borç başladı."),
        (7, "Tükeniş", 'Sistem son hamleyi yapıyor. Hiç kazanç vermez oldu, ne yaparsa yapsın oyuncu kaybeder. Para bitti, kredi tükendi, eve borç birikti. Oyuncu artık "bağımlı" — dur diyemez, oyun zaten kazanmış durumda.', "Nasıl bu hale geldim? Eve ne diyeceğim? Ama dursam kayıplar kalır... bir spin daha.", "Tükeniş Aşaması — bağımlılık literatüründe son dönemin klinik adı.", "Geri ödeme: %~5 | Tavan: bahsin 0.1×'i | Şans: yok", "Bu, ortalama bir maaşın katlarına ulaştı."),
    ]
    for i, isim, arka, kafa, bilim, roz, bag in asamalar:
        add(f"Sol panel Aşama {i}", "anlatici.html", f"A{i}_ISIM", isim)
        add(f"Sol panel Aşama {i}", "anlatici.html", f"A{i}_SAHNE_ARKASI", arka)
        add(f"Sol panel Aşama {i}", "anlatici.html", f"A{i}_OYUNCU_KAFASI", kafa)
        add(f"Sol panel Aşama {i}", "anlatici.html", f"A{i}_BILIM", bilim)
        add(f"Sol panel Aşama {i}", "anlatici.html", f"A{i}_ROZETLER", roz)
        add(f"Sol panel Aşama {i}", "anlatici.html", f"A{i}_BAKIYE_BAGLAM", bag)

    add("Şablon JS", "anlatici.html", "SABLON_SPIN_INFO", "kalan spin sonra aşama X başlayacak (Unity JSON ile dolar)")
    add("Şablon JS", "anlatici.html", "SABLON_SPIN_INFO_A7", "Bakiye tükenince simülasyon biter")
    add("Şablon JS", "anlatici.html", "SABLON_ASAMA_NUM", "AŞAMA X / 7")

    _yaz_xlsx(out_xlsx, rows)

    with open(out_csv, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["Bölüm", "Kaynak", "Kod", "Türkçe düz metin"],
            delimiter=";",
            quoting=csv.QUOTE_MINIMAL,
        )
        w.writeheader()
        w.writerows(rows)

    print("OK xlsx:", out_xlsx.resolve())
    print("OK csv :", out_csv.resolve())
    print("Satır (veri):", len(rows))


if __name__ == "__main__":
    main()
